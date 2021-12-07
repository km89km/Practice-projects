#! python3
# multiplicationTable.py - creates multiplication table from arguments
# provided in the command line. Practice project from chapter 13 of 'Automate
# the boring stuff' by Al Sweigart.

import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

if len(sys.argv) != 2:
    print('Usage - Int required for multiplication results.')
    sys.exit()

# The number provided in the command line will serve as the basis for the
# program.
n = int(sys.argv[1])

# create a new excel workbook.
wb = openpyxl.Workbook()
sheet = wb.active
# set the freeze panes so that the base numbers are always visible.
sheet.freeze_panes = 'B2'

# bold font for headers
bold = Font(bold=True)

# Create the base numbers along the top row and far left column.
# loop through the desired base numbers starting from 1.
for i in range(1, n + 1):
    # find the column and row to easily input the desired number.
    column = get_column_letter(i + 1)
    row = str(i + 1)
    # row of starting numbers
    sheet[column + '1'].font = bold
    sheet[column + '1'] = i

    # column of starting numbers
    sheet['A' + row].font = bold
    sheet['A' + row] = i

# set the bottom right cell to a variable to make it easier in the next section
# to get the desired multiplication results.
bottom_right = (get_column_letter(n + 1) + str(n + 1))

# loop through each row in the desired area.
for rowOfCellObjects in sheet['B2': bottom_right]:
    # loop through each cell in the current row.
    for cellObj in rowOfCellObjects:
        current_cell = cellObj.coordinate
        current_column = get_column_letter(cellObj.column)
        current_row = str(cellObj.row)
        # set the value of the current cell to the result of the multiplication.
        sheet[current_cell] = int(sheet['A' + current_row].value * (
            sheet[current_column + '1']).value)

# save the information to an excel file of the desired name.
wb.save(f'{n}xMultiplication.xlsx')
