#! python3
# cellInverter.py - inverts the column and row values of a specified excel file.
# I.e. cell 'D3' becomes 'C4' etc. Practice project from chapter 13 of
# 'Automate the Boring Stuff' by Al Sweigart.

import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# create new instance of workbook and select the current sheet.
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# read in the excel file and sheet that you want to invert.
old_wb = openpyxl.load_workbook('example.xlsx')
old_sheet = old_wb.active

# find the total area that is to be copied from the original file.
last_row = old_sheet.max_row
last_column = old_sheet.max_column
cut_off_cell = old_sheet[get_column_letter(last_column) +
                         str(last_row)].coordinate

# loop through each cell and write to new file.
for rowOfCellObjects in old_sheet['A1':cut_off_cell]:
    for cellObj in rowOfCellObjects:
        # to get new coordinate we get convert the row to the relevant column
        # letter and then add the str of the number of the column.
        new_coordinate = get_column_letter(cellObj.row) + str(cellObj.column)
        # we then set the value of this inverted coordinate to the
        # value from before.
        new_sheet[new_coordinate] = cellObj.value

# save the new file with inverted values.
new_wb.save('invertedcells.xlsx')
