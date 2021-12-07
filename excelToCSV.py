#! python3
# excelToCSV.py - converts all .xlsx files found in a provided
# directory to csv files.

import os
import openpyxl, csv
from openpyxl.utils import get_column_letter
from pathlib import Path

wf = # FOLDER TO BE USED

for excelFile in os.listdir(wf):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    else:
        wb = openpyxl.load_workbook(Path(wf, excelFile))
        for sheetName in wb.sheetnames:
            # Loop through every sheet in the workbook.
            sheet = wb[sheetName]

            # Create the CSV filename from the Excel filename adn sheet title.
            csvFilename = f"{excelFile.replace('.xlsx', '')}{sheet.title}.csv"
            # Create the csv.writer object for this CSV file.
            outputFile = open(csvFilename, 'w', newline='')
            outputWriter = csv.writer(outputFile)

            # Loop through every row in  the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []    # append each cell to this list
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    value = sheet[get_column_letter(colNum) + str(rowNum)].value
                    rowData.append(value)
                # Write the rowData list to the CSV file.
                outputWriter.writerow(rowData)
            outputFile.close()
