# ! python3
# blankRowInserter.py - Inserts the desired number of empty rows (N) at
# the indicated row (M) from the command line to a particular excel
# file. i.e 'blankRowInserter M N FILENAME.xlsx' Practice project from
# chapter 13 of 'Automate the boring stuff' by Al Sweigart'.

import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# check that sufficient arguments are provided in the command line.
if len(sys.argv) != 4:
    print('Usage - blankRowInserter M N filename.')
    sys.exit()

# save the provided arguments for insertion point and number of
# blank rows to variables.
insertion_row = int(sys.argv[1])
empty_rows = int(sys.argv[2])

# create new instance of Workbook to copy data to from specified file.
wbnew = openpyxl.Workbook()
new_sheet = wbnew.active

# open specified original excel file.
wbold = openpyxl.load_workbook(sys.argv[3])
old_sheet = wbold.active

# find total area before new blank rows to be copied.
last_row = old_sheet.max_row
last_column = get_column_letter(old_sheet.max_column)
cut_off_cell = old_sheet[last_column + str(insertion_row - 1)].coordinate

# using a for loop, copy each cell to new file.
for rowOfCellObjects in old_sheet['A1': cut_off_cell]:
    for cellObj in rowOfCellObjects:
        # As we simply copy it as is, we use the same cells as the old file.
        location = cellObj.coordinate
        new_sheet[location] = cellObj.value

# blank rows to be inserted.
finish_cell = str(insertion_row + (empty_rows - 1))
middle_start = new_sheet['A' + str(insertion_row)].coordinate
middle_finish = new_sheet[last_column + finish_cell].coordinate

# As this deals only with the new file, we just loop through the cells
# in the new file.
for rowOfCellObjects in new_sheet[middle_start: middle_finish]:
    for cellObj in rowOfCellObjects:
        cellObj.value = ''

# section after blank rows.
end_start = old_sheet['A' + str(insertion_row)].coordinate
end_finish = old_sheet[last_column + str(last_row)].coordinate
for rowOfCellObjects in old_sheet[end_start:end_finish]:
    for cellObj in rowOfCellObjects:
        # to get the coordinates of the cells where we want to copy the final
        # part, we add the number of empty rows to the row
        # coordinate from the old file.
        new_location = get_column_letter(cellObj.column) + str(
            cellObj.row + empty_rows)
        new_sheet[new_location].value = cellObj.value

# save the new file.
wbnew.save('newcells.xlsx')
