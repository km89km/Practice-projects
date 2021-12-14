#! python3
# textToSpreadsheet.py - reads in specified text files and writes each line to
# a cell in a column. I.e. the first file will be contained in column A and the
# second in column B and so on. Practice project from chapter 13 of
# 'Automate the Boring Stuff' by Al Sweigart.

import openpyxl
from openpyxl.utils import get_column_letter

# create a tuple containing all the files to be written to spreadsheet.
textFiles = ('text1.txt', 'text2.txt')

# create new instance of Workbook.
wb = openpyxl.Workbook()
sheet = wb.active

# iterate through each text file, using index to help with coordinates later.
for i in range(len(textFiles)):
    # open file, save lines to tuple and close file again.
    file = open(textFiles[i])
    lines = tuple(file.readlines())
    file.close()
    # iterate through the lines, again with index to help with coordinates.
    for j in range(len(lines)):
        # as the iterations start at 0, we add 1 to each as the rows and
        # columns of the cells start with 1.
        sheet[get_column_letter(i + 1) + str(j + 1)] = lines[j]

# save the new spreadsheet.
wb.save('textResult.xlsx')
